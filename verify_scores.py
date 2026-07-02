"""
verify_scores.py — Trail & Bish Dynasty Score Integrity Audit
==============================================================
Run during ANNUAL_UPDATE Step 5 (the audit pass):

    python verify_scores.py

What it checks
--------------
1. IMPLIED BASELINE DRIFT
   Computes each player's implied baseline:
       baseline = OVERALL SCORE − 0.6 × min(award_pts, 16)
   On first run, snapshots this to score_baseline_snapshot.json.
   On every subsequent run, diffs the current values against the snapshot.
   Baselines should NEVER move — any shift means an award count or score
   was edited without applying the matching delta. These are bugs.

2. NOTE / SB WIN CONSISTENCY
   Any player note containing "champion" or "Super Bowl" should have SB Win ≥ 1.
   Any player with SB Win ≥ 1 should have a note mentioning their ring.
   Flags both directions.

3. SCORE BOUNDS
   Every scored player must have 0.0 ≤ OVERALL SCORE ≤ 100.0.

4. PENDING COUNT
   Counts players with no OVERALL SCORE (the newest draft class).
   Reports the count and year so it's easy to confirm the right class is pending.

Award weights (LOCKED — matches scoring formula in CLAUDE.md):
    MVP=6, SB_MVP=4, OPOY=3, DPOY=3, ALL_PRO=2.5, SB Win=2.5, OROY=1, DROY=1
    Award cap = 16 pts before the 0.6× multiplier.
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT      = Path(__file__).resolve().parent
BACKEND   = ROOT / "backend" / "AP_Dynasty_Backend.xlsx"
SNAPSHOT  = ROOT / "score_baseline_snapshot.json"

# ── Award weights (LOCKED) ─────────────────────────────────────────────────
AWARD_WEIGHTS: dict[str, float] = {
    "MVP":     6.0,
    "SB_MVP":  4.0,
    "OPOY":    3.0,
    "DPOY":    3.0,
    "ALL_PRO": 2.5,
    "SB Win":  2.5,
    "OROY":    1.0,
    "DROY":    1.0,
}
AWARD_CAP   = 16.0
MULTIPLIER  = 0.6

# ── Ring note phrases ──────────────────────────────────────────────────────
_RING_WORDS = ("champion", "super bowl", "sb ", "ring")


# ── Helpers ────────────────────────────────────────────────────────────────

def _find_header(ws) -> tuple[int, dict[str, int]]:
    for rn in range(1, min(12, ws.max_row + 1)):
        vals = [str(c.value or "").strip() for c in ws[rn]]
        if "PLAYER" in vals and "OVERALL SCORE" in vals:
            return rn, {v: i + 1 for i, v in enumerate(vals) if v}
    raise ValueError("Could not find PLAYER / OVERALL SCORE header in 'players' sheet.")


def _award_pts(row: dict) -> float:
    pts = sum(
        float(row.get(col) or 0) * weight
        for col, weight in AWARD_WEIGHTS.items()
    )
    return min(pts, AWARD_CAP)


def _implied_baseline(score: float, award_pts: float) -> float:
    return round(score - MULTIPLIER * award_pts, 4)


def _note_claims_ring(note: str) -> bool:
    low = note.lower()
    return any(w in low for w in _RING_WORDS)


# ── Main ───────────────────────────────────────────────────────────────────

def main() -> None:
    print("Trail & Bish Dynasty — Score Integrity Audit")
    print(f"Date   : {date.today().isoformat()}")
    print(f"Source : {BACKEND.name}")
    print()

    if not BACKEND.exists():
        print(f"ERROR: backend not found at {BACKEND}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(BACKEND), data_only=True)
    ws = wb["players"]
    header_row, cols = _find_header(ws)

    col_player = cols["PLAYER"]
    col_score  = cols["OVERALL SCORE"]
    col_notes  = cols.get("NOTES")

    players: list[dict] = []
    for rn in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(rn, col_player).value
        if not name or not str(name).strip():
            continue
        row: dict = {"_row": rn}
        for col_name, col_idx in cols.items():
            row[col_name] = ws.cell(rn, col_idx).value
        players.append(row)

    total = len(players)
    issues: list[str] = []
    warnings: list[str] = []

    # ── 1. Implied baseline snapshot / diff ────────────────────────────────
    current_baselines: dict[str, float] = {}

    scored = [p for p in players if p.get("OVERALL SCORE") is not None]
    for p in scored:
        name  = str(p["PLAYER"]).strip()
        score = float(p["OVERALL SCORE"])
        apts  = _award_pts(p)
        base  = _implied_baseline(score, apts)
        current_baselines[name] = base

    first_run = not SNAPSHOT.exists()
    if first_run:
        SNAPSHOT.write_text(
            json.dumps(
                {"generated": date.today().isoformat(), "baselines": current_baselines},
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        print(f"[SNAPSHOT] Created {SNAPSHOT.name} with {len(current_baselines)} baselines.")
        print("           Run again after any data edit to diff against this snapshot.")
        print()
    else:
        snapshot_data = json.loads(SNAPSHOT.read_text(encoding="utf-8"))
        old_baselines: dict[str, float] = snapshot_data.get("baselines", {})
        snap_date = snapshot_data.get("generated", "unknown")

        drifted: list[tuple[str, float, float]] = []
        new_players: list[str] = []
        missing: list[str] = []

        for name, base in current_baselines.items():
            if name not in old_baselines:
                new_players.append(name)
            elif abs(base - old_baselines[name]) > 0.001:
                drifted.append((name, old_baselines[name], base))

        for name in old_baselines:
            if name not in current_baselines:
                missing.append(name)

        if drifted:
            for name, old, new in drifted:
                issues.append(
                    f"BASELINE DRIFT  {name}: was {old:.4f}, now {new:.4f}  "
                    f"(delta {new - old:+.4f}) — award count or score edited "
                    f"without matching delta"
                )
        if new_players:
            print(f"[SNAPSHOT] {len(new_players)} new scored player(s) since {snap_date}:")
            for n in new_players:
                print(f"           + {n}  (baseline: {current_baselines[n]:.4f})")
        if missing:
            warnings.append(
                f"{len(missing)} player(s) in snapshot no longer in workbook: "
                + ", ".join(missing)
            )

        if drifted or new_players or missing:
            # Update snapshot to include new players; keep existing baselines
            merged = {**old_baselines, **{n: current_baselines[n] for n in new_players}}
            SNAPSHOT.write_text(
                json.dumps(
                    {"generated": date.today().isoformat(), "baselines": merged},
                    indent=2,
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

    # ── 2. Note / SB Win consistency ───────────────────────────────────────
    if col_notes:
        for p in players:
            name     = str(p["PLAYER"]).strip()
            note     = str(p.get("NOTES") or "").strip()
            sb_wins  = int(p.get("SB Win") or 0)
            note_ring = _note_claims_ring(note) if note else False

            if note_ring and sb_wins == 0:
                issues.append(
                    f"NOTE/DATA MISMATCH  {name}: note mentions ring but SB Win = 0"
                )
            if sb_wins > 0 and note and not note_ring:
                warnings.append(
                    f"MISSING NOTE  {name}: SB Win = {sb_wins} but note has no ring mention"
                )
            if sb_wins > 0 and not note:
                warnings.append(
                    f"MISSING NOTE  {name}: SB Win = {sb_wins} but note is blank"
                )

    # ── 3. Score bounds ────────────────────────────────────────────────────
    for p in players:
        score = p.get("OVERALL SCORE")
        if score is None:
            continue
        s = float(score)
        if not (0.0 <= s <= 100.0):
            issues.append(
                f"SCORE OUT OF BOUNDS  {p['PLAYER']}: {s}"
            )

    # ── 4. Pending count ───────────────────────────────────────────────────
    pending = [p for p in players if p.get("OVERALL SCORE") is None]
    pending_years = sorted({int(p["YEAR"]) for p in pending if p.get("YEAR") is not None})

    # ── Report ─────────────────────────────────────────────────────────────
    print(f"Players     : {total} total  |  {len(scored)} scored  |  {len(pending)} pending")
    if pending_years:
        print(f"Pending year: {', '.join(str(y) for y in pending_years)}")
    print()

    if issues:
        print(f"{'='*60}")
        print(f"FAIL — {len(issues)} issue(s) require correction:")
        for i, msg in enumerate(issues, 1):
            print(f"  {i}. {msg}")
        print(f"{'='*60}")
    else:
        print("PASS — no data integrity issues found.")

    if warnings:
        print()
        print(f"WARNINGS ({len(warnings)}) — review but not necessarily bugs:")
        for w in warnings:
            print(f"  • {w}")

    print()
    if not first_run and not issues and not warnings:
        print("Audit clean. Snapshot is current.")
    elif not first_run and issues:
        print("Fix all FAIL items before committing. Re-run to confirm clean.")

    sys.exit(1 if issues else 0)


if __name__ == "__main__":
    main()
