"""
entry.py — Trail & Bish Dynasty Annual Draft-Entry Tool
=========================================================
Run locally once per year in April (Step 6 of ANNUAL_UPDATE.md):

    streamlit run entry.py

Parses a pasted prospect list into a searchable dropdown so no names are
ever free-typed.  Nothing touches the workbook until you hit Commit.
A backup is always created first.  A change-log is written to
draft_entry_{YEAR}.txt on commit.

Do NOT run this on Streamlit Cloud — it needs local write access to the workbook.
"""

from __future__ import annotations

import io
import re
import shutil
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).resolve().parent
BACKEND    = ROOT / "backend" / "AP_Dynasty_Backend.xlsx"
DRAFT_YEAR = date.today().year

POSITIONS = [
    "QB", "RB", "FB", "WR", "TE",
    "OT", "OG", "OC", "C", "OL",
    "DE", "DT", "NT", "EDGE", "ILB", "OLB", "LB", "DL",
    "CB", "S", "FS", "SS", "DB",
    "K", "P", "LS",
]
_POS_RE = re.compile(
    r'\b(' + '|'.join(POSITIONS) + r')\b',
    re.IGNORECASE,
)

# ── Page config + CSS ───────────────────────────────────────────────────────
st.set_page_config(
    page_title=f"Draft Entry {DRAFT_YEAR} · Trail & Bish",
    page_icon="🏈",
    layout="wide",
    initial_sidebar_state="collapsed",
)
st.markdown(
    """
    <style>
    html, body, [data-testid="stApp"] {
        background-color: #070B13; color: #C8D8F0;
        font-family: 'Segoe UI', system-ui, sans-serif;
    }
    [data-testid="stMainBlockContainer"] { padding-top: 1.5rem; }
    .hero {
        background: linear-gradient(135deg, #0D1421 0%, #0A1828 100%);
        border: 1px solid #1C2A45; border-radius: 12px;
        padding: 22px 30px; margin-bottom: 20px;
    }
    .phase-label {
        font-size: 10px; font-weight: 700; letter-spacing: .12em;
        text-transform: uppercase; color: #4A7BC4; margin-bottom: 6px;
    }
    .pick-row {
        display: flex; align-items: center; gap: 8px;
        padding: 5px 10px; border-radius: 6px;
        background: #0D1421; border: 1px solid #1C2A45;
        margin-bottom: 4px; font-size: 13px;
    }
    .cat-badge {
        font-size: 10px; font-weight: 700; letter-spacing: .06em;
        background: #1C2A45; color: #7BA8D4; border-radius: 4px;
        padding: 2px 7px;
    }
    .commit-log {
        background: #060f08; border: 1px solid #1a4020; border-radius: 8px;
        padding: 14px 18px; font-family: 'Courier New', monospace;
        font-size: 12px; color: #6fcf8a; white-space: pre-wrap;
        max-height: 360px; overflow-y: auto; margin-top: 12px;
    }
    .stButton > button { border-radius: 6px; font-weight: 600; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Session state ────────────────────────────────────────────────────────────
for _k, _v in [
    ("prospects", []),  # list[(name, pos)]
    ("staged",    []),  # list[{player, position, owner, category}]
    ("committed", False),
    ("commit_log", ""),
    ("parse_done", False),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ── Helpers: parsing ─────────────────────────────────────────────────────────

def parse_prospects(text: str) -> list[tuple[str, str]]:
    """
    Extract (name, position) pairs from free-form pasted text.
    Handles: "1. Cam Ward, QB, Miami" / "Cam Ward QB" / "1) Cam Ward QB" etc.
    """
    results: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        # Strip leading rank: "1. " / "1) " / "1 - " / "1 "
        line = re.sub(r'^[\d]+[\.\)\-:\s]\s*', '', line).strip()
        m = _POS_RE.search(line)
        if not m:
            continue
        pos  = m.group(1).upper()
        name = line[:m.start()].strip().rstrip(',-/ ').strip()
        name = re.sub(r'\s+', ' ', name)
        # Drop stray trailing words that aren't part of the name (e.g. school)
        name = re.split(r',', name)[0].strip()
        if len(name) < 3 or name.upper() in ('NAME', 'PLAYER'):
            continue
        key = (name.lower(), pos)
        if key not in seen:
            seen.add(key)
            results.append((name, pos))
    return results


# ── Helpers: workbook ────────────────────────────────────────────────────────

def workbook_writable() -> bool:
    try:
        with open(BACKEND, 'r+b'):
            return True
    except (FileNotFoundError, PermissionError, OSError):
        return False


def _find_players_header(ws) -> tuple[int, dict[str, int]]:
    for rn in range(1, min(12, ws.max_row + 1)):
        vals = [str(c.value or '').strip() for c in ws[rn]]
        if 'PLAYER' in vals and 'OVERALL SCORE' in vals:
            return rn, {v: i + 1 for i, v in enumerate(vals) if v}
    raise ValueError("Could not find PLAYER / OVERALL SCORE header in 'players' sheet.")


def _last_data_row(ws, col: int = 3) -> int:
    """Last row index where the given column has a non-blank value."""
    last = 0
    for rn in range(1, ws.max_row + 1):
        v = ws.cell(rn, col).value
        if v is not None and str(v).strip():
            last = rn
    return last


def commit_picks(staged: list[dict], year: int) -> tuple[bool, str, str]:
    """
    Write staged picks to the workbook.
    Returns (success, log_text, error_message).
    Always backs up first.  Uses write_bytes() to avoid NTFS truncation.
    """
    if not BACKEND.exists():
        return False, '', f"Workbook not found: {BACKEND}"

    # Backup (don't overwrite an existing backup from the same day)
    bak = BACKEND.with_stem(BACKEND.stem + f"_bak_{date.today():%Y%m%d}")
    if not bak.exists():
        shutil.copy2(BACKEND, bak)

    wb = openpyxl.load_workbook(str(BACKEND))

    main_picks = [p for p in staged if p['category'] == 'Main']
    wc_picks   = [p for p in staged if p['category'] != 'Main']

    log_lines: list[str] = [
        "Trail & Bish Dynasty — Draft Entry Log",
        f"Date: {date.today():%Y-%m-%d}  |  Year: {year}  |  Backup: {bak.name}",
        "",
    ]

    # ── players sheet ────────────────────────────────────────────────────────
    ws_p = wb['players']
    _, cols = _find_players_header(ws_p)

    c_player = cols['PLAYER']
    c_pos    = cols.get('POSITION', 2)
    c_owner  = cols.get('OWNER', 3)
    c_year   = cols.get('YEAR', 4)

    start_row = ws_p.max_row + 1
    log_lines.append(f"players sheet — {len(main_picks)} new rows (starting at row {start_row})")

    matt_names = [p['player'] for p in main_picks if p['owner'] == 'Matt']
    ryan_names = [p['player'] for p in main_picks if p['owner'] == 'Ryan']
    log_lines.append(f"  Matt ({len(matt_names)}): {', '.join(matt_names) or '—'}")
    log_lines.append(f"  Ryan ({len(ryan_names)}): {', '.join(ryan_names) or '—'}")

    for p in main_picks:
        ws_p.cell(start_row, c_player).value = p['player']
        ws_p.cell(start_row, c_pos).value    = p['position']
        ws_p.cell(start_row, c_owner).value  = p['owner']
        ws_p.cell(start_row, c_year).value   = year
        # All award columns, OVERALL SCORE, NOTES: intentionally left blank (TBD / pending)
        start_row += 1

    # ── wildcard boys sheet ─────────────────────────────────────────────────
    if wc_picks:
        ws_wc     = wb['wildcard boys']
        wc_start  = _last_data_row(ws_wc, col=3) + 1
        log_lines.append(f"\nwildcard boys sheet — {len(wc_picks)} new rows (starting at row {wc_start})")
        for p in wc_picks:
            ws_wc.cell(wc_start, 1).value = year           # YEAR
            ws_wc.cell(wc_start, 2).value = p['owner']     # OWNER
            ws_wc.cell(wc_start, 3).value = p['player']    # PLAYER
            ws_wc.cell(wc_start, 4).value = p['position']  # POSITION
            ws_wc.cell(wc_start, 5).value = p['category']  # CATEGORY (WC1/WC2/WC3)
            # OUTCOME, COOKED_METER, NOTES: blank (Pending)
            log_lines.append(
                f"  {p['owner']} {p['category']}: {p['player']} ({p['position']})"
            )
            wc_start += 1

    # Save using write_bytes to avoid NTFS write-buffer truncation
    buf = io.BytesIO()
    wb.save(buf)
    BACKEND.write_bytes(buf.getvalue())

    # Write log file
    log_text  = '\n'.join(log_lines)
    log_path  = ROOT / f"draft_entry_{year}.txt"
    log_path.write_text(log_text, encoding='utf-8')

    return True, log_text, ''


# ═══════════════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════════════

# Hero
st.markdown(
    f"""
    <div class="hero">
        <div style="font-size:22px;font-weight:800;letter-spacing:-.01em">
            🏈 Draft Entry &nbsp;·&nbsp; {DRAFT_YEAR}
        </div>
        <div style="color:#4A7BC4;margin-top:4px;font-size:13px">
            Trail &amp; Bish Dynasty &nbsp;·&nbsp;
            ANNUAL_UPDATE Step 6 &nbsp;·&nbsp;
            <strong>Run locally only</strong> — Streamlit Cloud cannot write to the workbook
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# Environment gate
if not BACKEND.exists():
    st.error(
        f"Workbook not found at `{BACKEND}`.  \n"
        "Run this from the project root: `streamlit run entry.py`"
    )
    st.stop()

if not workbook_writable():
    st.error(
        "**The workbook is not writable from here.**  \n"
        "This tool must run locally on your own machine — not on Streamlit Cloud.  \n"
        "Also make sure `AP_Dynasty_Backend.xlsx` isn't open in Excel."
    )
    st.stop()

# ─── PHASE 1: Paste & Parse ─────────────────────────────────────────────────
st.markdown('<div class="phase-label">Phase 1 — Paste Prospect List</div>', unsafe_allow_html=True)

with st.expander(
    "Paste your prospect list here",
    expanded=not st.session_state.parse_done,
):
    raw_text = st.text_area(
        label="Copy from PFF, The Draft Network, ESPN, or any plain-text source.",
        height=230,
        placeholder=(
            "1. Cam Ward, QB, Miami\n"
            "2. Travis Hunter, WR, Colorado\n"
            "3. Shedeur Sanders, QB, Colorado\n"
            "…\n"
            "Handles numbered lists, bullet points, and simple 'Name Pos' lines."
        ),
        key="paste_text",
    )
    c1, c2 = st.columns([2, 5])
    if c1.button("Parse List →", type="primary", use_container_width=True):
        if raw_text.strip():
            parsed = parse_prospects(raw_text)
            if parsed:
                st.session_state.prospects  = parsed
                st.session_state.parse_done = True
                st.rerun()
            else:
                st.error(
                    "No name + position pairs found. "
                    "Make sure each line has a recognisable position "
                    "(QB, WR, OT, EDGE, etc.)."
                )
        else:
            st.error("Paste a prospect list first.")
    c2.caption(
        "Flexible parser: strips rank numbers, handles commas, "
        "Name/Pos/School and similar formats."
    )

if st.session_state.parse_done:
    n = len(st.session_state.prospects)
    st.success(f"✓ {n} prospects parsed")

    with st.expander(f"View parsed list ({n})", expanded=False):
        st.dataframe(
            pd.DataFrame(st.session_state.prospects, columns=["Name", "Position"]),
            use_container_width=True,
            height=min(280, 40 + n * 35),
            hide_index=True,
        )
    if st.button("↩ Re-paste / re-parse", use_container_width=False):
        st.session_state.parse_done = False
        st.session_state.prospects  = []
        st.rerun()

if not st.session_state.parse_done:
    st.stop()

st.divider()

# ─── PHASE 2: Build the Class ───────────────────────────────────────────────
st.markdown('<div class="phase-label">Phase 2 — Build the Class</div>', unsafe_allow_html=True)

prospects     = st.session_state.prospects
display_opts  = [f"{n}  ({p})" for n, p in prospects] + ["— manual entry —"]
prospect_map  = {f"{n}  ({p})": (n, p) for n, p in prospects}

col_matt, col_ryan = st.columns(2)

for owner, col, color, emoji in [
    ("Matt", col_matt, "#2E7DF7", "🔵"),
    ("Ryan", col_ryan, "#E63B3B", "🔴"),
]:
    with col:
        owner_picks = [p for p in st.session_state.staged if p['owner'] == owner]
        main_n = sum(1 for p in owner_picks if p['category'] == 'Main')
        wc_n   = sum(1 for p in owner_picks if p['category'] != 'Main')

        st.markdown(
            f'<div style="color:{color};font-size:17px;font-weight:800;'
            f'margin-bottom:6px">{emoji} {owner}'
            f'<span style="color:#4A7BC4;font-size:12px;font-weight:400;'
            f'margin-left:10px">{main_n} main · {wc_n} wildcard</span></div>',
            unsafe_allow_html=True,
        )

        # Pick entry form
        with st.form(key=f"form_{owner}", clear_on_submit=True):
            chosen = st.selectbox(
                "Select prospect",
                options=display_opts,
                key=f"sel_{owner}",
            )
            manual_name = st.text_input(
                "Manual name (only if not in list above)",
                key=f"man_{owner}",
                placeholder="Leave blank to use dropdown",
            )
            manual_pos = st.selectbox(
                "Manual position",
                options=["(use dropdown)"] + POSITIONS,
                key=f"manpos_{owner}",
            )
            category = st.radio(
                "Category",
                options=["Main", "WC1", "WC2", "WC3"],
                horizontal=True,
                key=f"cat_{owner}",
            )
            submitted = st.form_submit_button(
                f"Add to {owner}'s {DRAFT_YEAR} class",
                use_container_width=True,
            )
            if submitted:
                if manual_name.strip():
                    p_name = manual_name.strip()
                    p_pos  = (
                        manual_pos
                        if manual_pos != "(use dropdown)"
                        else "?"
                    )
                elif chosen == "— manual entry —":
                    st.warning("Select a prospect or enter a name manually.")
                    p_name, p_pos = None, None
                else:
                    p_name, p_pos = prospect_map[chosen]

                if p_name:
                    st.session_state.staged.append({
                        "player":   p_name,
                        "position": p_pos,
                        "owner":    owner,
                        "category": category,
                    })
                    st.rerun()

        # Staged picks for this owner
        if owner_picks:
            st.markdown("**Staged:**")
            global_indices = [
                gi for gi, sp in enumerate(st.session_state.staged)
                if sp['owner'] == owner
            ]
            for gi in global_indices:
                p = st.session_state.staged[gi]
                c_pick, c_del = st.columns([6, 1])
                c_pick.markdown(
                    f'<div class="pick-row">'
                    f'<span class="cat-badge">{p["category"]}</span>'
                    f'<span style="flex:1">{p["player"]}</span>'
                    f'<span style="color:#4A7BC4;font-size:12px">{p["position"]}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                if c_del.button("✕", key=f"del_{gi}"):
                    st.session_state.staged.pop(gi)
                    st.rerun()

st.divider()

# ─── PHASE 3: Review & Commit ───────────────────────────────────────────────
st.markdown('<div class="phase-label">Phase 3 — Review &amp; Commit</div>', unsafe_allow_html=True)

staged = st.session_state.staged

if not staged:
    st.info("No picks staged yet — add picks above.")
    st.stop()

main_n = sum(1 for p in staged if p['category'] == 'Main')
wc_n   = sum(1 for p in staged if p['category'] != 'Main')
st.caption(f"{len(staged)} total picks — {main_n} main / {wc_n} wildcard")

st.dataframe(
    pd.DataFrame(staged)[["owner", "category", "player", "position"]].rename(
        columns={"owner": "Owner", "category": "Category",
                 "player": "Player", "position": "Pos"}
    ),
    use_container_width=True,
    hide_index=True,
    height=min(400, 60 + len(staged) * 36),
)

if st.session_state.committed:
    st.success(
        f"✓ Committed — {len(staged)} picks written to "
        f"`backend/AP_Dynasty_Backend.xlsx`. "
        f"Log saved to `draft_entry_{DRAFT_YEAR}.txt`."
    )
    st.markdown(
        f'<div class="commit-log">{st.session_state.commit_log}</div>',
        unsafe_allow_html=True,
    )
    st.info(
        f"**Next steps:**  \n"
        f"1. `python make_premium.py` — regenerate the Premium workbook with the new class.  \n"
        f"2. `git add backend/AP_Dynasty_Backend.xlsx Trail_Bish_Dynasty_Premium.xlsx "
        f"draft_entry_{DRAFT_YEAR}.txt`  \n"
        f"3. `git commit -m \"D0XX: {DRAFT_YEAR} draft class entered — "
        f"{main_n} main picks, {wc_n} wildcards\"`  \n"
        f"4. Push → Streamlit Cloud redeploys.  \n"
        f"5. Verify sidebar shows the new player count."
    )
else:
    c_commit, c_clear = st.columns([3, 1])
    if c_commit.button(
        f"🏈  Commit {len(staged)} picks to workbook",
        type="primary",
        use_container_width=True,
    ):
        with st.spinner("Writing to workbook…"):
            ok, log_text, err = commit_picks(staged, DRAFT_YEAR)
        if ok:
            st.session_state.committed  = True
            st.session_state.commit_log = log_text
            st.rerun()
        else:
            st.error(f"Commit failed: {err}")

    if c_clear.button("Clear all picks", use_container_width=True):
        st.session_state.staged = []
        st.rerun()

    st.caption(
        "⚠ This will write to `AP_Dynasty_Backend.xlsx` directly. "
        "A dated backup is created automatically before any write."
    )
