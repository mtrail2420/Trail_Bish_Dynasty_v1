"""
make_premium.py — Trail & Bish Dynasty Premium Workbook Generator
==================================================================
Reads AP_Dynasty_Backend.xlsx and rebuilds Trail_Bish_Dynasty_Premium.xlsx
from scratch with dark-navy styling.  Run after every data edit:

    python make_premium.py

The Premium workbook is OUTPUT, not source.  Never edit it directly.
All corrections go to AP_Dynasty_Backend.xlsx first, then regenerate.

Sheets produced:
  1. DYNASTY          — title / intro card
  2. PLAYERS          — full 354-row roster with dark navy formatting
  3. MAN STATUS       — head-to-head rivalry scorecard
  4. WILDCARD BOYS    — wildcard picks + Cooked Meter
  5. SCORING MODEL    — scoring formula reference

Color palette (locked):
  Background   #070B13     dark navy
  Surface      #0D1421     card surface
  Border       #1C2A45     subtle border
  Matt         #2E7DF7     blue
  Ryan         #E63B3B     red
  Gold         #D4AF37
  Text primary #C8D8F0
  Text muted   #4A7BC4
"""

from __future__ import annotations

import io
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT    = Path(__file__).resolve().parent
BACKEND = ROOT / "backend" / "AP_Dynasty_Backend.xlsx"
OUTPUT  = ROOT / "Trail_Bish_Dynasty_Premium.xlsx"

# ── Color palette (RGB hex, no #) ─────────────────────────────────────────
BG      = "070B13"   # darkest navy
SURFACE = "0D1421"   # card surface
BORDER  = "1C2A45"   # subtle border
MATT    = "2E7DF7"   # Matt blue
RYAN    = "E63B3B"   # Ryan red
GOLD    = "D4AF37"   # gold accent
TEXT    = "C8D8F0"   # primary text
MUTED   = "4A7BC4"   # muted text / labels
GREEN   = "22C55E"   # positive / win
WHITE   = "FFFFFF"


# ── Style helpers ──────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(
    hex_color: str = TEXT,
    bold: bool = False,
    size: int = 10,
    name: str = "Calibri",
) -> Font:
    return Font(color=hex_color, bold=bold, size=size, name=name)


def _border(hex_color: str = BORDER) -> Border:
    side = Side(style="thin", color=hex_color)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_cell(
    ws,
    row: int,
    col: int,
    value,
    *,
    bg: str = BG,
    fg: str = TEXT,
    bold: bool = False,
    size: int = 10,
    h: str = "left",
    v: str = "center",
    wrap: bool = False,
    border: bool = False,
    num_format: str | None = None,
) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(fg, bold=bold, size=size)
    c.alignment = _align(h, v, wrap)
    if border:
        c.border = _border()
    if num_format:
        c.number_format = num_format


def _header_row(ws, row: int, headers: list[tuple[str, int, str]]) -> None:
    """
    headers: list of (label, col_index, hex_color)
    """
    for label, col, color in headers:
        _set_cell(
            ws, row, col, label,
            bg=SURFACE, fg=color, bold=True, size=9,
            h="center", border=True,
        )


def _fill_row(ws, row: int, ncols: int, bg: str) -> None:
    """Paint background across a row range."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
            cell.fill = _fill(bg)


def _col_widths(ws, widths: dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


# ── Tier logic (mirrors core/stats.py) ────────────────────────────────────
TIER_MAP = [
    (95.0, "Legend",             "D4AF37"),
    (80.0, "Franchise",          "2E7DF7"),
    (68.0, "High-End Starter",   "22C55E"),
    (54.0, "Starter",            "C8D8F0"),
    (40.0, "Contributor",        "4A7BC4"),
    (0.0,  "Bust",               "E63B3B"),
]

def score_to_tier(score) -> tuple[str, str]:
    """Return (tier_label, hex_color)."""
    if score is None:
        return "Pending", "4A7BC4"
    s = float(score)
    for threshold, label, color in TIER_MAP:
        if s >= threshold:
            return label, color
    return "Bust", "E63B3B"


# ── Backend reader ─────────────────────────────────────────────────────────

def _find_header(ws, required=("PLAYER", "OVERALL SCORE")):
    for rn in range(1, min(12, ws.max_row + 1)):
        vals = [str(c.value or "").strip() for c in ws[rn]]
        if all(r in vals for r in required):
            return rn, {v: i + 1 for i, v in enumerate(vals) if v}
    raise ValueError(f"Could not find header with {required}")


def _read_players(wb) -> tuple[int, dict[str, int], list[dict]]:
    ws = wb["players"]
    header_row, cols = _find_header(ws, ("PLAYER", "OVERALL SCORE"))
    rows = []
    for rn in range(header_row + 1, ws.max_row + 1):
        player = ws.cell(rn, cols["PLAYER"]).value
        if not player or not str(player).strip():
            continue
        row = {}
        for col_name, col_idx in cols.items():
            row[col_name] = ws.cell(rn, col_idx).value
        rows.append(row)
    return header_row, cols, rows


def _read_sheet_raw(wb, sheet_name: str):
    """Return list of row-dicts from a sheet (openpyxl raw)."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


# ══════════════════════════════════════════════════════════════════════════
# Sheet builders
# ══════════════════════════════════════════════════════════════════════════

def _build_dynasty(wb_out):
    ws = wb_out.active
    ws.title = "DYNASTY"

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    # column widths
    widths = {1: 4, 2: 36, 3: 26, 4: 22, 5: 18, 6: 4}
    _col_widths(ws, widths)

    def _bg(row: int, ncols: int = 6, color: str = BG) -> None:
        for c in range(1, ncols + 1):
            ws.cell(row, c).fill = _fill(color)

    # Paint all rows dark
    for r in range(1, 42):
        _bg(r)

    # Row 2 — main title
    _row_height(ws, 2, 60)
    ws.merge_cells("B2:E2")
    _set_cell(ws, 2, 2, "TRAIL & BISH DYNASTY", bg=BG, fg=GOLD, bold=True, size=28, h="center")

    # Row 4 — subtitle
    _row_height(ws, 4, 22)
    ws.merge_cells("B4:E4")
    _set_cell(ws, 4, 2, "Two Friends. Twenty Draft Classes. One Rivalry.", bg=BG, fg=MUTED, bold=False, size=13, h="center")

    # Row 6 — divider label
    _row_height(ws, 6, 18)
    ws.merge_cells("B6:E6")
    _set_cell(ws, 6, 2, "THE WORKBOOK", bg=BG, fg=MUTED, bold=True, size=9, h="center")

    # Row 8-10 — stat cards (filled in after we load data)
    return ws


def _fill_dynasty_stats(ws, players: list[dict]) -> None:
    """
    Populate stat cards after players data is loaded.
    """
    matt_players = [p for p in players if str(p.get("OWNER") or "").strip().lower() == "matt"]
    ryan_players = [p for p in players if str(p.get("OWNER") or "").strip().lower() == "ryan"]

    def _avg(rows):
        scores = [float(p["OVERALL SCORE"]) for p in rows if p.get("OVERALL SCORE") is not None]
        return round(sum(scores) / len(scores), 1) if scores else 0.0

    def _franchise(rows):
        return sum(1 for p in rows if p.get("OVERALL SCORE") is not None and float(p["OVERALL SCORE"]) >= 80)

    stats = [
        ("TOTAL PLAYERS", str(len(players)), ""),
        ("MATT'S ROSTER", str(len(matt_players)), f"avg {_avg(matt_players):.1f}"),
        ("RYAN'S ROSTER", str(len(ryan_players)), f"avg {_avg(ryan_players):.1f}"),
        ("DRAFT CLASSES", "20", "2007 – 2026"),
        ("FRANCHISE PLAYERS", str(_franchise(players)), "score ≥ 80"),
    ]

    start_row = 8
    _row_height(ws, start_row, 18)
    _row_height(ws, start_row + 1, 30)
    _row_height(ws, start_row + 2, 14)

    col_map = [2, 3, 4, 5]
    colors  = [MUTED, MATT, RYAN, GOLD, GREEN]

    for i, (label, value, sub) in enumerate(stats):
        col = col_map[i] if i < len(col_map) else 2
        color = colors[i] if i < len(colors) else TEXT
        _set_cell(ws, start_row,     col, label, bg=SURFACE, fg=MUTED, bold=True, size=8, h="center")
        _set_cell(ws, start_row + 1, col, value, bg=SURFACE, fg=color,  bold=True, size=22, h="center")
        _set_cell(ws, start_row + 2, col, sub,   bg=SURFACE, fg=MUTED, bold=False, size=8, h="center")

    # 5th stat (Franchise) into row 12
    _row_height(ws, 12, 18)
    _row_height(ws, 13, 30)
    _row_height(ws, 14, 14)
    label, value, sub = stats[4]
    _set_cell(ws, 12, 2, label, bg=SURFACE, fg=MUTED,  bold=True,  size=8,  h="center")
    _set_cell(ws, 13, 2, value, bg=SURFACE, fg=GOLD,   bold=True,  size=22, h="center")
    _set_cell(ws, 14, 2, sub,   bg=SURFACE, fg=MUTED,  bold=False, size=8,  h="center")


def _build_players(wb_out, players: list[dict]) -> None:
    ws = wb_out.create_sheet("PLAYERS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = MATT

    # ── Columns: RANK PLAYER POSITION OWNER YEAR SCORE TIER NOTES
    COLS = [
        ("RANK",     5),
        ("PLAYER",   30),
        ("POSITION", 10),
        ("OWNER",    10),
        ("YEAR",      7),
        ("SCORE",     9),
        ("TIER",      18),
        ("NOTES",     50),
    ]
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title rows
    _row_height(ws, 1, 36)
    _row_height(ws, 2, 18)
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    _set_cell(ws, 1, 1, "PLAYERS", bg=BG, fg=TEXT, bold=True, size=20, h="center")
    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")
    _set_cell(ws, 2, 1, "Full Dynasty Roster — All 354 Picks", bg=BG, fg=MUTED, bold=False, size=10, h="center")

    # Header row
    HEADER_ROW = 3
    _row_height(ws, HEADER_ROW, 18)
    header_labels = [(label, i + 1, MUTED) for i, (label, _) in enumerate(COLS)]
    _header_row(ws, HEADER_ROW, header_labels)

    # Sort: by OVERALL SCORE descending
    sorted_players = sorted(
        players,
        key=lambda p: float(p.get("OVERALL SCORE") or 0),
        reverse=True,
    )

    for rank, p in enumerate(sorted_players, 1):
        rn = HEADER_ROW + rank
        _row_height(ws, rn, 15)

        score = p.get("OVERALL SCORE")
        tier_label, tier_color = score_to_tier(score)
        owner = str(p.get("OWNER") or "").strip()
        owner_color = MATT if owner.lower() == "matt" else (RYAN if owner.lower() == "ryan" else TEXT)
        row_bg = SURFACE if rank % 2 == 0 else BG

        score_str = f"{float(score):.1f}" if score is not None else "—"

        vals = [
            (rank,                    TEXT,       "center"),
            (p.get("PLAYER") or "",   TEXT,       "left"),
            (p.get("POSITION") or "", MUTED,      "center"),
            (owner,                   owner_color,"center"),
            (p.get("YEAR") or "",     MUTED,      "center"),
            (score_str,               tier_color, "center"),
            (tier_label,              tier_color, "center"),
            (p.get("NOTES") or "",    MUTED,      "left"),
        ]

        for col_idx, (val, fg, h) in enumerate(vals, 1):
            _set_cell(ws, rn, col_idx, val, bg=row_bg, fg=fg, h=h,
                      border=(col_idx != len(COLS)),  # no border on NOTES for readability
                      wrap=(col_idx == len(COLS)))

    # Freeze panes below header
    ws.freeze_panes = f"A{HEADER_ROW + 1}"


def _build_man_status(wb_out, players: list[dict]) -> None:
    ws = wb_out.create_sheet("MAN STATUS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RYAN

    for i, w in enumerate([4, 22, 14, 14, 14, 14, 14, 14, 4], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _row_height(ws, 1, 44)
    _row_height(ws, 2, 20)
    _row_height(ws, 3, 30)
    _row_height(ws, 4, 18)

    ws.merge_cells("B1:H1")
    _set_cell(ws, 1, 2, "MAN STATUS", bg=BG, fg=TEXT, bold=True, size=26, h="center")
    ws.merge_cells("B2:H2")
    _set_cell(ws, 2, 2, "The Rivalry. The Record. The Legend.", bg=BG, fg=MUTED, size=11, h="center")

    # Header row 3
    headers = [
        (3, "MATT",      MATT),
        (4, "RYAN",      RYAN),
        (5, "ADVANTAGE", GOLD),
        (6, "STAT",      MUTED),
        (7, "MATT",      MATT),
        (8, "RYAN",      RYAN),
    ]
    for col, label, color in headers:
        _set_cell(ws, 3, col, label, bg=SURFACE, fg=color, bold=True, size=10, h="center", border=True)

    # Compute rivalry stats
    matt_p = [p for p in players if str(p.get("OWNER") or "").strip().lower() == "matt"]
    ryan_p = [p for p in players if str(p.get("OWNER") or "").strip().lower() == "ryan"]

    def _scores(rows):
        return [float(p["OVERALL SCORE"]) for p in rows if p.get("OVERALL SCORE") is not None]

    def _avg(rows):
        s = _scores(rows)
        return round(sum(s) / len(s), 2) if s else 0.0

    def _tier_count(rows, lo, hi=200.0):
        return sum(1 for p in rows if p.get("OVERALL SCORE") is not None and lo <= float(p["OVERALL SCORE"]) < hi)

    def _award_sum(rows, col):
        return sum(int(p.get(col) or 0) for p in rows)

    rivalry_rows = [
        ("Total Picks",     len(matt_p),                               len(ryan_p)),
        ("Avg Score",       _avg(matt_p),                              _avg(ryan_p)),
        ("Franchise (80+)", _tier_count(matt_p, 80),                   _tier_count(ryan_p, 80)),
        ("HES (68–79.9)",   _tier_count(matt_p, 68, 80),               _tier_count(ryan_p, 68, 80)),
        ("Starters (54–67.9)", _tier_count(matt_p, 54, 68),            _tier_count(ryan_p, 54, 68)),
        ("Busts (<40)",     _tier_count(matt_p, 0, 40),                 _tier_count(ryan_p, 0, 40)),
        ("MVP Awards",      _award_sum(matt_p, "MVP"),                  _award_sum(ryan_p, "MVP")),
        ("SB Wins",         _award_sum(matt_p, "SB Win"),               _award_sum(ryan_p, "SB Win")),
        ("All-Pro (total)", _award_sum(matt_p, "ALL_PRO"),              _award_sum(ryan_p, "ALL_PRO")),
        ("SB MVP",          _award_sum(matt_p, "SB_MVP"),               _award_sum(ryan_p, "SB_MVP")),
        ("OPOY",            _award_sum(matt_p, "OPOY"),                 _award_sum(ryan_p, "OPOY")),
        ("DPOY",            _award_sum(matt_p, "DPOY"),                 _award_sum(ryan_p, "DPOY")),
    ]

    for i, (label, m_val, r_val) in enumerate(rivalry_rows):
        rn  = 4 + i
        _row_height(ws, rn, 16)
        row_bg = SURFACE if i % 2 == 0 else BG

        if isinstance(m_val, float):
            m_str = f"{m_val:.2f}"
            r_str = f"{r_val:.2f}"
            m_gt  = m_val > r_val
        else:
            m_str = str(m_val)
            r_str = str(r_val)
            m_gt  = m_val > r_val

        adv_label = "MATT" if m_gt else ("RYAN" if r_val > m_val else "TIED")
        adv_color = MATT if m_gt else (RYAN if r_val > m_val else MUTED)

        _set_cell(ws, rn, 2, label,     bg=row_bg, fg=TEXT,  bold=False, h="left",   border=True)
        _set_cell(ws, rn, 3, m_str,     bg=row_bg, fg=MATT,  bold=True,  h="center", border=True)
        _set_cell(ws, rn, 4, r_str,     bg=row_bg, fg=RYAN,  bold=True,  h="center", border=True)
        _set_cell(ws, rn, 5, adv_label, bg=row_bg, fg=adv_color, bold=True, h="center", border=True)

    ws.freeze_panes = "B5"


def _build_wildcard_boys(wb_out, wb_backend) -> None:
    ws = wb_out.create_sheet("WILDCARD BOYS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN

    # Read raw from backend
    if "wildcard boys" not in wb_backend.sheetnames:
        ws.cell(1, 1).value = "wildcard boys sheet not found in backend"
        return

    ws_src = wb_backend["wildcard boys"]
    raw_rows = list(ws_src.iter_rows(values_only=True))

    col_widths = {1: 6, 2: 10, 3: 26, 4: 12, 5: 14, 6: 14, 7: 12, 8: 40}
    _col_widths(ws, col_widths)

    # Title
    _row_height(ws, 1, 36)
    ws.merge_cells("A1:H1")
    _set_cell(ws, 1, 1, "WILDCARD BOYS", bg=BG, fg=GREEN, bold=True, size=22, h="center")

    # Find data start — wildcard boys sheet has data from row 8 (index 7)
    DATA_START_IDX = 7  # 0-indexed
    header_labels = ["YEAR", "OWNER", "PLAYER", "POSITION", "CATEGORY", "OUTCOME", "COOKED METER", "NOTES"]

    _row_height(ws, 2, 18)
    for ci, label in enumerate(header_labels, 1):
        _set_cell(ws, 2, ci, label, bg=SURFACE, fg=MUTED, bold=True, size=9, h="center", border=True)

    out_row = 3
    for ri, row in enumerate(raw_rows):
        if ri < DATA_START_IDX:
            continue
        if not any(v for v in row):
            continue

        owner = str(row[1] or "").strip() if len(row) > 1 else ""
        owner_color = MATT if owner.lower() == "matt" else (RYAN if owner.lower() == "ryan" else TEXT)
        row_bg = SURFACE if out_row % 2 == 0 else BG
        _row_height(ws, out_row, 15)

        values = list(row)[:8] + [None] * max(0, 8 - len(row))
        colors = [MUTED, owner_color, TEXT, MUTED, MUTED, MUTED, GOLD, MUTED]
        haligns = ["center", "center", "left", "center", "center", "center", "center", "left"]

        for ci, (val, fg, ha) in enumerate(zip(values, colors, haligns), 1):
            _set_cell(ws, out_row, ci, val, bg=row_bg, fg=fg, h=ha, border=True, wrap=(ci == 8))

        out_row += 1

    ws.freeze_panes = "A3"


def _build_scoring_model(wb_out) -> None:
    ws = wb_out.create_sheet("SCORING MODEL")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    col_widths = {1: 4, 2: 28, 3: 14, 4: 14, 5: 4}
    _col_widths(ws, col_widths)

    for r in range(1, 60):
        for c in range(1, 6):
            ws.cell(r, c).fill = _fill(BG)

    _row_height(ws, 1, 36)
    ws.merge_cells("B1:D1")
    _set_cell(ws, 1, 2, "SCORING MODEL", bg=BG, fg=GOLD, bold=True, size=22, h="center")

    _row_height(ws, 2, 18)
    ws.merge_cells("B2:D2")
    _set_cell(ws, 2, 2, "Formula reference — read-only", bg=BG, fg=MUTED, size=9, h="center")

    # Formula
    _row_height(ws, 4, 20)
    ws.merge_cells("B4:D4")
    _set_cell(ws, 4, 2, "OVERALL SCORE  =  min(100, baseline + 0.6 × min(award_pts, 16))",
              bg=SURFACE, fg=TEXT, bold=True, size=11, h="center")

    # Award weights table
    award_weights = [
        ("Award",              "Weight", "Points"),
        ("MVP",                6,        "× 0.6 = 3.6 pts"),
        ("SB MVP",             4,        "× 0.6 = 2.4 pts"),
        ("OPOY / DPOY",        3,        "× 0.6 = 1.8 pts"),
        ("All-Pro / SB Win",   2.5,      "× 0.6 = 1.5 pts"),
        ("OROY / DROY",        1,        "× 0.6 = 0.6 pts"),
        ("Award cap",          16,       "max award pts"),
    ]

    start = 6
    for i, (a, w, pts) in enumerate(award_weights):
        rn = start + i
        _row_height(ws, rn, 16)
        is_header = (i == 0)
        row_bg = SURFACE if is_header or i % 2 == 1 else BG
        fg = GOLD if is_header else TEXT
        _set_cell(ws, rn, 2, a,   bg=row_bg, fg=fg, bold=is_header, h="left",   border=True)
        _set_cell(ws, rn, 3, w,   bg=row_bg, fg=MUTED if not is_header else fg, bold=is_header, h="center", border=True)
        _set_cell(ws, rn, 4, pts, bg=row_bg, fg=MUTED if not is_header else fg, bold=is_header, h="left",   border=True)

    # Tier thresholds
    tier_start = start + len(award_weights) + 2
    ws.merge_cells(f"B{tier_start}:D{tier_start}")
    _set_cell(ws, tier_start, 2, "TIER THRESHOLDS", bg=SURFACE, fg=GOLD, bold=True, size=11, h="center")

    tiers = [
        ("Legend",           "≥ 95.0",  GOLD),
        ("Franchise",        "80 – 94.9", MATT),
        ("High-End Starter", "68 – 79.9", GREEN),
        ("Starter",          "54 – 67.9", TEXT),
        ("Contributor",      "40 – 53.9", MUTED),
        ("Bust",             "< 40",     RYAN),
    ]
    for j, (tier, rng, color) in enumerate(tiers):
        rn = tier_start + 1 + j
        _row_height(ws, rn, 16)
        row_bg = SURFACE if j % 2 == 0 else BG
        _set_cell(ws, rn, 2, tier, bg=row_bg, fg=color, bold=True, h="left",   border=True)
        _set_cell(ws, rn, 3, rng,  bg=row_bg, fg=MUTED, bold=False, h="center", border=True)


# ══════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("Trail & Bish Dynasty — Premium Workbook Generator")
    print(f"Date: {date.today().isoformat()}")
    print(f"Source: {BACKEND}")
    print(f"Output: {OUTPUT}")
    print()

    if not BACKEND.exists():
        print(f"ERROR: backend not found at {BACKEND}")
        print("Run from the project root: python make_premium.py")
        return

    # Backup existing Premium if it exists
    if OUTPUT.exists():
        bak = OUTPUT.with_stem(OUTPUT.stem + f"_bak_{date.today():%Y%m%d}")
        if not bak.exists():
            shutil.copy2(OUTPUT, bak)
            print(f"Backed up → {bak.name}")

    print("Reading backend…")
    wb_backend = openpyxl.load_workbook(str(BACKEND))
    _, _, players = _read_players(wb_backend)
    print(f"  {len(players)} players loaded")

    print("Building Premium workbook…")
    wb_out = openpyxl.Workbook()

    # Sheet 1: DYNASTY
    _build_dynasty(wb_out)
    _fill_dynasty_stats(wb_out.active, players)
    print("  ✓ DYNASTY")

    # Sheet 2: PLAYERS
    _build_players(wb_out, players)
    print("  ✓ PLAYERS")

    # Sheet 3: MAN STATUS
    _build_man_status(wb_out, players)
    print("  ✓ MAN STATUS")

    # Sheet 4: WILDCARD BOYS
    _build_wildcard_boys(wb_out, wb_backend)
    print("  ✓ WILDCARD BOYS")

    # Sheet 5: SCORING MODEL
    _build_scoring_model(wb_out)
    print("  ✓ SCORING MODEL")

    # Global workbook properties
    wb_out.properties.title    = "Trail & Bish Dynasty Premium"
    wb_out.properties.subject  = f"Generated {date.today().isoformat()}"
    wb_out.properties.creator  = "make_premium.py"
    wb_out.active              = wb_out["DYNASTY"]

    # Save using write_bytes to avoid NTFS truncation
    print(f"\nSaving → {OUTPUT.name}…")
    buf = io.BytesIO()
    wb_out.save(buf)
    OUTPUT.write_bytes(buf.getvalue())

    size_kb = OUTPUT.stat().st_size / 1024
    print(f"Done — {size_kb:.1f} KB")
    print()
    print("Next steps:")
    print("  git add Trail_Bish_Dynasty_Premium.xlsx")
    print(f"  git commit -m 'chore: regenerate Premium ({date.today():%Y-%m-%d})'")
    print("  git push → Streamlit Cloud redeploys")


if __name__ == "__main__":
    main()
