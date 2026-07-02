"""
fix_audit.py  —  Trail & Bish Dynasty full award + notes audit
Run this ONCE from the project root on a machine where the xlsx file is readable:

    python fix_audit.py

What it does:
  • Backs up AP_Dynasty_Backend.xlsx before touching anything
  • Applies 9 award / score corrections to the players sheet
  • Removes the obsolete "Franchise designation removed" sentence from 11 player notes
  • Updates Kenneth Walker III's note (SB LX champion + SB LX MVP)
  • Checks / updates JSN, Ernest Jones IV, Charles Cross notes for SB LX
  • Prints a summary of every change made

After running this, regenerate the Premium workbook:
    python make_premium.py

Rule: corrections go to the backend only.
      The Premium is output — never edit it directly.
"""

from __future__ import annotations
import re
import shutil
from datetime import date
from pathlib import Path

import openpyxl

# ── Paths ───────────────────────────────────────────────────────────────────
ROOT        = Path(__file__).resolve().parent
BACKEND     = ROOT / "backend" / "AP_Dynasty_Backend.xlsx"
BACKUP_DATE = date.today().strftime("%Y%m%d")

# ── Award / Score corrections (source: real-world audit) ────────────────────
#
# Keys are exact player names as they appear in the workbook.
# col_deltas: column-header → new absolute value (None = no change).
# new_score  : new OVERALL SCORE value.
# note       : human-readable reason (logged only).
#
AWARD_FIXES: list[dict] = [
    {
        "player":    "Jameis Winston",
        "col_deltas": {"SB Win": 0},
        "new_score":  60.0,
        "reason":    "Never on a champion roster — SB Win removed",
    },
    {
        "player":    "Chase Winovich",
        "col_deltas": {"SB Win": 0},
        "new_score":  28.0,
        "reason":    "Retired 2023, no ring — SB Win removed",
    },
    {
        "player":    "Jawaan Taylor",
        "col_deltas": {"SB Win": 1},
        "new_score":  69.5,
        "reason":    "Only KC LVIII (not LVII) — SB Win 2→1",
    },
    {
        "player":    "Myles Garrett",
        "col_deltas": {"ALL_PRO": 6},
        "new_score":  97.6,          # unchanged (already over award cap)
        "reason":    "2025 first-team AP unanimous — ALL_PRO 5→6, score unchanged (over cap)",
    },
    {
        "player":    "Cam Newton",
        "col_deltas": {"ALL_PRO": 1},
        "new_score":  92.5,
        "reason":    "2015 first-team AP All-Pro (MVP season) — ALL_PRO 0→1",
    },
    {
        "player":    "Roquan Smith",
        "col_deltas": {"ALL_PRO": 2},
        "new_score":  81.0,
        "reason":    "Audit confirms 2 first-team APs (2022, 2023), not 3 — ALL_PRO 3→2",
    },
    {
        "player":    "DeVonta Smith",
        "col_deltas": {"SB Win": 1},
        "new_score":  67.5,
        "reason":    "Eagles SB LIX champion — SB Win 0→1",
    },
    {
        "player":    "Mekhi Becton",
        "col_deltas": {"SB Win": 1},
        "new_score":  59.5,
        "reason":    "Eagles starting RG, SB LIX — SB Win 0→1",
    },
    {
        "player":    "Leonard Williams",
        "col_deltas": {"SB Win": 1},
        "new_score":  75.5,
        "reason":    "Seahawks SB LX roster — SB Win 0→1",
    },
]

# ── Sentence to remove from 11 player notes ─────────────────────────────────
OBSOLETE_SENTENCE = (
    "Franchise designation removed during Legacy Audit; "
    "career projection not counted as earned franchise status."
)
NOTES_CLEANUP_PLAYERS = [
    "Kenneth Walker III",
    "Tyler Linderbaum",
    "Jalen Carter",
    "Brian Branch",
    "Jared Verse",
    "Jayden Daniels",
    "Brock Bowers",
    "Joe Alt",
    "Abdul Carter",
    "Ashton Jeanty",
    "Mason Graham",
]

# ── Kenneth Walker III update ────────────────────────────────────────────────
WALKER_SB_ADDITION = "Super Bowl LX champion and Super Bowl LX MVP."

# ── Players needing SB LX note if not already present ───────────────────────
SB_LX_PLAYERS = ["Jaxon Smith-Njigba", "Ernest Jones", "Ernest Jones IV", "Charles Cross"]
SB_LX_PHRASE  = "Super Bowl LX champion"

# ── Helpers ─────────────────────────────────────────────────────────────────

def backup(path: Path) -> Path:
    dst = path.with_stem(path.stem + f"_bak_{BACKUP_DATE}")
    shutil.copy2(path, dst)
    print(f"  Backup → {dst.name}")
    return dst


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    """
    Scan the sheet for the header row and return (row_number, {header: col_index}).
    col_index is 1-based.
    """
    for row in ws.iter_rows(max_row=10):
        vals = [str(c.value).strip() if c.value is not None else "" for c in row]
        if "PLAYER" in vals and "OVERALL SCORE" in vals:
            return row[0].row, {v: i + 1 for i, v in enumerate(vals)}
    raise RuntimeError("Could not find header row in sheet (expected PLAYER, OVERALL SCORE, ...)")


def find_player_row(ws, col_player: int, name: str) -> int | None:
    """Return the 1-based row number for the given player name, or None."""
    for row in ws.iter_rows(min_col=col_player, max_col=col_player):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == name.strip().lower():
            return cell.row
    return None


def clean_sentence(note: str | None, sentence: str) -> str:
    """Remove a sentence from a note string, cleaning up extra whitespace."""
    if not note:
        return note or ""
    cleaned = note.replace(sentence, "").strip()
    # Collapse double spaces / leading-trailing spaces
    cleaned = re.sub(r"  +", " ", cleaned).strip()
    return cleaned


def append_to_note(note: str | None, addition: str) -> str:
    """Append `addition` to an existing note, guarding against duplicates."""
    note = (note or "").strip()
    if addition.lower() in note.lower():
        return note  # already there
    if note and not note.endswith("."):
        note += "."
    return (note + " " + addition).strip() if note else addition


# ── Main ─────────────────────────────────────────────────────────────────────

def apply_fixes(wb_path: Path, is_backend: bool) -> None:
    label = "backend" if is_backend else "premium"
    print(f"\n{'='*60}")
    print(f"Processing {label}: {wb_path.name}")
    print(f"{'='*60}")

    if not wb_path.exists():
        print(f"  SKIP — file not found: {wb_path}")
        return

    backup(wb_path)

    wb = openpyxl.load_workbook(str(wb_path))

    # ── Locate the players sheet ─────────────────────────────────────────────
    players_sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == "players":
            players_sheet_name = name
            break
    if players_sheet_name is None:
        print("  ERROR: no 'players' sheet found")
        return

    ws = wb[players_sheet_name]
    header_row, cols = find_header_row(ws)
    print(f"  Players sheet: '{players_sheet_name}', header at row {header_row}, {ws.max_row - header_row} data rows")

    col_player = cols.get("PLAYER")
    col_score  = cols.get("OVERALL SCORE")
    col_notes  = cols.get("NOTES") if is_backend else None  # premium may not have NOTES

    if not col_player or not col_score:
        print("  ERROR: required columns PLAYER / OVERALL SCORE not found")
        return

    changes: list[str] = []

    # ── 1. Award + score corrections ─────────────────────────────────────────
    print("\n  --- Award / Score Corrections ---")
    for fix in AWARD_FIXES:
        pname = fix["player"]
        rn = find_player_row(ws, col_player, pname)
        if rn is None:
            print(f"  NOT FOUND: {pname}")
            continue

        made_change = False

        # Update award columns
        for award_col_name, new_val in fix["col_deltas"].items():
            col_idx = cols.get(award_col_name)
            if col_idx is None:
                continue
            old_val = ws.cell(rn, col_idx).value
            if old_val != new_val:
                ws.cell(rn, col_idx).value = new_val
                changes.append(f"{pname} [{award_col_name}] {old_val!r} → {new_val!r}")
                made_change = True

        # Update score
        old_score = ws.cell(rn, col_score).value
        if old_score != fix["new_score"]:
            ws.cell(rn, col_score).value = fix["new_score"]
            changes.append(f"{pname} [OVERALL SCORE] {old_score!r} → {fix['new_score']!r}")
            made_change = True

        status = "✓ CHANGED" if made_change else "  no-op (already correct)"
        print(f"  {status}: {pname}  — {fix['reason']}")

    # ── 2. Notes cleanup (backend only — Premium has no NOTES column) ─────────
    if is_backend and col_notes:
        print("\n  --- Notes Cleanup (remove obsolete sentence) ---")
        for pname in NOTES_CLEANUP_PLAYERS:
            rn = find_player_row(ws, col_player, pname)
            if rn is None:
                print(f"  NOT FOUND: {pname}")
                continue
            old_note = ws.cell(rn, col_notes).value or ""
            if OBSOLETE_SENTENCE in old_note:
                ws.cell(rn, col_notes).value = clean_sentence(old_note, OBSOLETE_SENTENCE)
                changes.append(f"{pname} [NOTES] — obsolete sentence removed")
                print(f"  ✓ Cleaned: {pname}")
            else:
                print(f"    already clean: {pname}")

    # ── 3. Kenneth Walker III — add SB LX info ───────────────────────────────
    if is_backend and col_notes:
        print("\n  --- Kenneth Walker III Note Update ---")
        rn = find_player_row(ws, col_player, "Kenneth Walker III")
        if rn:
            old_note = ws.cell(rn, col_notes).value or ""
            new_note = append_to_note(old_note, WALKER_SB_ADDITION)
            if new_note != old_note:
                ws.cell(rn, col_notes).value = new_note
                changes.append("Kenneth Walker III [NOTES] — SB LX champion + MVP added")
                print(f"  ✓ Updated Kenneth Walker III")
            else:
                print(f"    already up to date")
        else:
            print("  NOT FOUND: Kenneth Walker III")

    # ── 4. JSN / Ernest Jones / Charles Cross — SB LX check ─────────────────
    if is_backend and col_notes:
        print("\n  --- SB LX Notes Check (JSN, Ernest Jones, Charles Cross) ---")
        for pname in SB_LX_PLAYERS:
            rn = find_player_row(ws, col_player, pname)
            if rn is None:
                continue  # try all variants
            old_note = ws.cell(rn, col_notes).value or ""
            if SB_LX_PHRASE.lower() in old_note.lower():
                print(f"    already has SB LX: {pname}")
            else:
                ws.cell(rn, col_notes).value = append_to_note(old_note, SB_LX_PHRASE + ".")
                changes.append(f"{pname} [NOTES] — SB LX champion added")
                print(f"  ✓ Updated: {pname}")

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(str(wb_path))
    print(f"\n  Saved → {wb_path.name}")
    print(f"  Total changes: {len(changes)}")
    for c in changes:
        print(f"    • {c}")


def main() -> None:
    print("Trail & Bish Dynasty — Full Award + Notes Audit Fix")
    print(f"Date: {date.today().isoformat()}\n")

    if not BACKEND.exists():
        print(f"ERROR: backend not found at {BACKEND}")
        return

    apply_fixes(BACKEND, is_backend=True)

    print("\n" + "="*60)
    print("DONE.  Next steps:")
    print("  1. Open AP_Dynasty_Backend.xlsx in Excel, verify changes.")
    print("  2. python make_premium.py   ← regenerate Premium from updated backend")
    print("  3. git add backend/AP_Dynasty_Backend.xlsx Trail_Bish_Dynasty_Premium.xlsx")
    print("  4. git commit -m 'D036: full award + notes audit — 9 fixes, 11 cleanups'")
    print("  5. git push → Streamlit Cloud redeploys automatically.")
    print("  6. Check sidebar Data Status: should show ✓ 354 players · 3 sheets.")
    print("="*60)


if __name__ == "__main__":
    main()
